"""
TJRJ Precatórios Scraper V5 - All Entities In-Memory Mode

Key Features:
- Processes ALL entities in a single run (no per-entity I/O)
- Full in-memory accumulation until final save
- Entities processed in order (largest first for optimal worker usage)
- Screenshot capture on navigation failures for debugging
- Aggressive timeout handling on final pages
- Guaranteed browser cleanup even on errors

Usage:
    python main_v5_all_entities.py --regime especial --num-processes 12
    python main_v5_all_entities.py --regime geral --num-processes 8
"""

import sys
import time
import argparse
import multiprocessing as mp
import re
import signal
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from loguru import logger
import pandas as pd

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.scraper_v3 import TJRJPrecatoriosScraperV3
from src.models import ScraperConfig, EntidadeDevedora


# Global flag for graceful shutdown
SHUTDOWN_REQUESTED = False


def signal_handler(signum, frame):
    """Handle shutdown signals gracefully"""
    global SHUTDOWN_REQUESTED
    SHUTDOWN_REQUESTED = True
    logger.warning("⚠️ Shutdown requested - finishing current work...")


def setup_logging(level: str = "INFO", log_file: str = None):
    """Configure logging with timestamped log file"""
    logger.remove()
    
    # Console output
    logger.add(
        sys.stderr,
        format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <7}</level> | <level>{message}</level>",
        level=level,
        colorize=True
    )
    
    # File output - always write to scraper_v3.log for UI compatibility
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    # Main log file for UI parsing
    main_log = log_dir / "scraper_v3.log"
    logger.add(
        main_log,
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <7} | {message}",
        level="DEBUG",
        rotation="50 MB"
    )
    
    # Additional timestamped log if specified
    if log_file:
        log_path = Path(log_file)
        logger.add(
            log_path,
            format="{time:YYYY-MM-DD HH:mm:ss} | {level: <7} | {message}",
            level="DEBUG"
        )
        return str(log_path)
    
    return str(main_log)


def divide_pages_into_ranges(total_pages: int, num_processes: int) -> List[Tuple[int, int]]:
    """Divide total pages into ranges for parallel processing"""
    if total_pages == 0:
        return []
    
    # Don't create more workers than pages
    effective_processes = min(num_processes, total_pages)
    
    pages_per_process = total_pages // effective_processes
    remainder = total_pages % effective_processes
    
    ranges = []
    start = 1
    
    for i in range(effective_processes):
        extra = 1 if i < remainder else 0
        end = start + pages_per_process - 1 + extra
        ranges.append((start, end))
        start = end + 1
    
    return ranges


def take_debug_screenshot(page, process_id: int, context: str) -> Optional[str]:
    """Take a screenshot for debugging and return the path"""
    try:
        screenshot_dir = Path("logs/screenshots")
        screenshot_dir.mkdir(parents=True, exist_ok=True)
        
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        filename = f"P{process_id}_{context}_{timestamp}.png"
        filepath = screenshot_dir / filename
        
        page.screenshot(path=str(filepath), full_page=False)
        logger.info(f"[P{process_id}] 📸 Screenshot saved: {filepath}")
        return str(filepath)
    except Exception as e:
        logger.warning(f"[P{process_id}] ⚠️ Could not take screenshot: {e}")
        return None


def extract_worker(args: Dict) -> Dict:
    """
    Worker function for parallel extraction
    
    Returns dict with results directly in memory (NO disk I/O)
    Includes screenshot capture on failures
    """
    from playwright.sync_api import sync_playwright
    
    entity_id = args['entity_id']
    entity_name = args['entity_name']
    regime = args['regime']
    start_page = args['start_page']
    end_page = args['end_page']
    process_id = args['process_id']
    skip_expanded = args.get('skip_expanded', True)
    headless = args.get('headless', True)
    timeout_minutes = args.get('timeout_minutes', 30)
    
    start_time = time.time()
    timeout_seconds = timeout_minutes * 60
    
    # Accumulate records in memory only
    precatorios_data = []
    browser = None
    context = None
    
    try:
        logger.info(f"[P{process_id}] 🚀 Starting: pages {start_page}-{end_page} (timeout: {timeout_minutes}min)")
        
        # Create scraper
        config = ScraperConfig(headless=headless)
        scraper = TJRJPrecatoriosScraperV3(config=config, skip_expanded=skip_expanded)
        
        # Create entity
        entidade = EntidadeDevedora(
            id_entidade=entity_id,
            nome_entidade=entity_name,
            regime=regime,
            precatorios_pagos=0,
            precatorios_pendentes=0,
            valor_prioridade=0,
            valor_rpv=0
        )
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            context = browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/120.0.0.0 Safari/537.36'
            )
            page = context.new_page()
            
            # Set shorter default timeout for navigation
            page.set_default_timeout(30000)
            
            # Navigate to entity
            logger.info(f"[P{process_id}] 🌐 Navigating to entity page...")
            entity_url = f"https://www3.tjrj.jus.br/PortalConhecimento/precatorio#!/ordem-cronologica?idEntidadeDevedora={entity_id}"
            page.goto(entity_url, wait_until='networkidle', timeout=60000)
            page.wait_for_timeout(3000)
            logger.info(f"[P{process_id}] ✅ Entity page loaded")
            
            # Go to start page
            if start_page > 1:
                logger.info(f"[P{process_id}] 🔄 Jumping to start page {start_page}...")
                if not scraper.goto_page_direct(page, start_page):
                    take_debug_screenshot(page, process_id, f"nav_fail_page{start_page}")
                    raise Exception(f"Failed to navigate to page {start_page}")
                logger.info(f"[P{process_id}] ✅ Arrived at page {start_page}")
            
            # Extract pages
            current_page = start_page
            total_pages_in_range = end_page - start_page + 1
            consecutive_failures = 0
            max_consecutive_failures = 5
            
            while current_page <= end_page:
                # Check timeout
                elapsed = time.time() - start_time
                if elapsed > timeout_seconds:
                    logger.warning(f"[P{process_id}] ⏰ Timeout after {elapsed/60:.1f}min - saving {len(precatorios_data)} records")
                    take_debug_screenshot(page, process_id, f"timeout_page{current_page}")
                    break
                
                # Check shutdown flag
                if SHUTDOWN_REQUESTED:
                    logger.warning(f"[P{process_id}] 🛑 Shutdown requested - saving {len(precatorios_data)} records")
                    break
                
                page_in_range = current_page - start_page + 1
                
                # Heartbeat logging every 50 pages
                if page_in_range % 50 == 0:
                    speed = len(precatorios_data) / elapsed if elapsed > 0 else 0
                    logger.info(f"[P{process_id}] 💓 Heartbeat: {page_in_range}/{total_pages_in_range} pages, {len(precatorios_data)} records, {speed:.1f} rec/s")
                
                # Log format compatible with UI: [P1] Page X/Y (Z/W)
                logger.info(f"[P{process_id}] Page {current_page}/{end_page} ({page_in_range}/{total_pages_in_range})")
                
                try:
                    # Extract precatórios from current page
                    precatorios = scraper._extract_precatorios_from_page(page, entidade)
                    
                    # Convert to dicts and accumulate in memory
                    for prec in precatorios:
                        precatorios_data.append(prec.model_dump())
                    
                    # Log format compatible with UI: [P1] ✅ ... (total: N)
                    logger.info(f"[P{process_id}]   ✅ {len(precatorios)} records (total: {len(precatorios_data)})")
                    consecutive_failures = 0  # Reset on success
                    
                except Exception as extract_err:
                    logger.warning(f"[P{process_id}] ⚠️ Extraction error on page {current_page}: {extract_err}")
                    consecutive_failures += 1
                    
                    if consecutive_failures >= max_consecutive_failures:
                        logger.error(f"[P{process_id}] ❌ Too many consecutive failures - stopping")
                        take_debug_screenshot(page, process_id, f"consecutive_fail_page{current_page}")
                        break
                
                # Next page navigation
                if current_page < end_page:
                    pages_remaining = end_page - current_page
                    
                    # More aggressive handling for final pages (last 10)
                    if pages_remaining <= 10:
                        nav_success = False
                        for retry in range(3):
                            try:
                                # Wait for any loading overlay to disappear
                                try:
                                    page.wait_for_selector('.block-ui-overlay', state='hidden', timeout=2000)
                                except:
                                    pass
                                
                                next_btn = page.query_selector('a[ng-click="vm.ProximaPagina()"]')
                                if next_btn:
                                    # Scroll to button first (may be below viewport)
                                    try:
                                        next_btn.scroll_into_view_if_needed()
                                        page.wait_for_timeout(300)
                                    except:
                                        pass
                                    
                                    # Try to click
                                    try:
                                        next_btn.click()
                                        page.wait_for_timeout(2000)
                                        nav_success = True
                                        break
                                    except Exception as click_err:
                                        logger.warning(f"[P{process_id}] ⚠️ Click failed (retry {retry+1}): {click_err}")
                                        page.wait_for_timeout(1000)
                                else:
                                    # Button not found - take screenshot
                                    if retry == 2:
                                        take_debug_screenshot(page, process_id, f"no_next_btn_page{current_page}")
                                    page.wait_for_timeout(1000)
                            except Exception as nav_err:
                                logger.warning(f"[P{process_id}] ⚠️ Navigation error (retry {retry+1}): {nav_err}")
                                page.wait_for_timeout(1000)
                        
                        if not nav_success:
                            logger.warning(f"[P{process_id}] ⚠️ Failed to navigate after 3 retries on page {current_page}")
                            # STOP - don't continue with wrong page data
                            logger.warning(f"[P{process_id}] 🛑 Stopping worker to avoid duplicate data")
                            break
                    else:
                        # Fast navigation for most pages - also with scroll
                        try:
                            next_btn = page.query_selector('a[ng-click="vm.ProximaPagina()"]')
                            if next_btn:
                                try:
                                    next_btn.scroll_into_view_if_needed()
                                    page.wait_for_timeout(200)
                                except:
                                    pass
                                next_btn.click()
                                page.wait_for_timeout(2000)
                        except Exception as nav_err:
                            logger.warning(f"[P{process_id}] ⚠️ Navigation error: {nav_err}")
                
                current_page += 1
            
            # Close browser explicitly (like V4)
            logger.info(f"[P{process_id}] 🔄 Extraction complete, closing browser...")
            try:
                context.close()
                browser.close()
                logger.info(f"[P{process_id}] ✅ Browser closed")
            except Exception as close_error:
                logger.warning(f"[P{process_id}] ⚠️ Browser close issue: {close_error}")
        
        elapsed = time.time() - start_time
        logger.info(f"[P{process_id}] ✅ Complete: {len(precatorios_data)} records in {elapsed/60:.1f}min")
        
        # Return data in memory only - NO disk I/O
        return {
            'process_id': process_id,
            'entity_id': entity_id,
            'entity_name': entity_name,
            'start_page': start_page,
            'end_page': end_page,
            'records': precatorios_data,
            'records_count': len(precatorios_data),
            'elapsed_seconds': elapsed,
            'success': True,
            'error': None
        }
    
    except Exception as e:
        elapsed = time.time() - start_time
        logger.error(f"[P{process_id}] ❌ Failed: {e}")
        
        # Return whatever we have in memory
        return {
            'process_id': process_id,
            'entity_id': entity_id,
            'entity_name': entity_name,
            'start_page': start_page,
            'end_page': end_page,
            'records': precatorios_data,
            'records_count': len(precatorios_data),
            'elapsed_seconds': elapsed,
            'success': False,
            'error': str(e)
        }


def extract_single_entity(
    entity_id: int,
    entity_name: str,
    regime: str,
    total_pages: int,
    num_processes: int,
    headless: bool = True,
    timeout_minutes: int = 30
) -> Tuple[List[Dict], Dict]:
    """
    Extract all records from a single entity using parallel workers
    
    Returns:
        Tuple of (list of record dicts, stats dict)
    """
    if total_pages == 0:
        logger.info(f"⏭️ Skipping {entity_name} - no pages")
        return [], {'total_records': 0, 'success': True}
    
    # Divide pages among workers
    ranges = divide_pages_into_ranges(total_pages, num_processes)
    effective_workers = len(ranges)
    
    logger.info(f"\n{'='*80}")
    logger.info(f"🏛️ ENTITY: {entity_name} (ID: {entity_id})")
    logger.info(f"{'='*80}")
    logger.info(f"Pages: {total_pages} | Workers: {effective_workers} | Timeout: {timeout_minutes}min")
    
    for i, (start, end) in enumerate(ranges, 1):
        logger.info(f"  P{i}: pages {start:,}-{end:,} ({end-start+1:,} pages)")
    
    # Prepare worker args
    worker_args = []
    for i, (start, end) in enumerate(ranges, 1):
        worker_args.append({
            'entity_id': entity_id,
            'entity_name': entity_name,
            'regime': regime,
            'start_page': start,
            'end_page': end,
            'process_id': i,
            'skip_expanded': True,
            'headless': headless,
            'timeout_minutes': timeout_minutes
        })
    
    # Accumulate all records in memory
    all_records = []
    results = []
    
    worker_timeout = timeout_minutes * 60 + 60  # Add 1 min buffer
    
    logger.info(f"\n🔄 Starting extraction...")
    start_time = time.time()
    
    with mp.Pool(processes=effective_workers) as pool:
        # Submit all workers
        async_results = []
        for args in worker_args:
            async_results.append((args['process_id'], pool.apply_async(extract_worker, (args,))))
        
        # Collect results with timeout
        for process_id, async_result in async_results:
            try:
                result = async_result.get(timeout=worker_timeout)
                results.append(result)
                
                if result['success']:
                    all_records.extend(result['records'])
                    logger.info(f"✅ P{result['process_id']} done: {result['records_count']} records")
                else:
                    logger.error(f"❌ P{result['process_id']} failed: {result['error']}")
                    # Still add partial records
                    all_records.extend(result['records'])
                    
            except mp.TimeoutError:
                logger.error(f"❌ P{process_id} TIMEOUT after {worker_timeout}s")
                results.append({
                    'process_id': process_id,
                    'records': [],
                    'records_count': 0,
                    'success': False,
                    'error': f'Timeout after {worker_timeout}s'
                })
            except Exception as e:
                logger.error(f"❌ P{process_id} ERROR: {e}")
                results.append({
                    'process_id': process_id,
                    'records': [],
                    'records_count': 0,
                    'success': False,
                    'error': str(e)
                })
    
    elapsed = time.time() - start_time
    
    stats = {
        'entity_id': entity_id,
        'entity_name': entity_name,
        'total_records': len(all_records),
        'elapsed_seconds': elapsed,
        'successful_workers': sum(1 for r in results if r['success']),
        'failed_workers': sum(1 for r in results if not r['success']),
        'success': any(r['success'] for r in results)
    }
    
    logger.info(f"📊 Entity complete: {len(all_records)} records in {elapsed/60:.1f}min")
    
    return all_records, stats


def load_entities_from_website(regime: str, headless: bool = True) -> List[Dict]:
    """Load entity list from TJRJ website"""
    from playwright.sync_api import sync_playwright
    
    logger.info(f"📋 Loading entities for regime: {regime}")
    
    if regime == 'geral':
        url = "https://www3.tjrj.jus.br/PortalConhecimento/precatorio/#!/entes-devedores/regime-geral"
    else:
        url = "https://www3.tjrj.jus.br/PortalConhecimento/precatorio/#!/entes-devedores/regime-especial"
    
    entities = []
    
    with sync_playwright() as p:
        browser = p.chromium.launch(headless=headless)
        context = browser.new_context(
            viewport={'width': 1920, 'height': 1080},
            user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                      'AppleWebKit/537.36 (KHTML, like Gecko) '
                      'Chrome/120.0.0.0 Safari/537.36'
        )
        page = context.new_page()
        
        page.goto(url, wait_until='networkidle', timeout=60000)
        
        try:
            page.wait_for_selector("text=Precatórios Pagos", timeout=15000)
        except:
            logger.warning("Timeout waiting for entity cards")
        
        page.wait_for_timeout(3000)
        
        # Find entity cards
        cards = page.query_selector_all('[ng-repeat="ente in vm.EntesDevedores"]')
        logger.info(f"Found {len(cards)} entity cards")
        
        for card in cards:
            try:
                # Get card text and parse
                card_text = card.inner_text()
                lines = [line.strip() for line in card_text.split('\n') if line.strip()]
                
                # Get entity name (first line)
                name = lines[0] if lines else ''
                if not name:
                    continue
                
                # Get link for ID - try href first, then ng-click
                link = card.query_selector('a[href*="idEntidadeDevedora"]')
                if link:
                    href = link.get_attribute('href') or ''
                    id_match = re.search(r'idEntidadeDevedora=(\d+)', href)
                else:
                    link = card.query_selector('a[ng-click*="idEntidadeDevedora"]')
                    if link:
                        ng_click = link.get_attribute('ng-click') or ''
                        id_match = re.search(r'idEntidadeDevedora=(\d+)', ng_click)
                    else:
                        continue
                
                if not id_match:
                    continue
                
                entity_id = int(id_match.group(1))
                
                # Parse statistics from card text
                pendentes = 0
                pagos = 0
                
                for i, line in enumerate(lines):
                    if 'Precatórios Pagos' in line:
                        if i + 1 < len(lines):
                            num = re.sub(r'[^\d]', '', lines[i + 1])
                            pagos = int(num) if num else 0
                    
                    if 'Precatórios Pendentes' in line:
                        if i + 1 < len(lines):
                            num = re.sub(r'[^\d]', '', lines[i + 1])
                            pendentes = int(num) if num else 0
                
                entities.append({
                    'id': entity_id,
                    'nome': name,
                    'precatorios_pendentes': pendentes,
                    'precatorios_pagos': pagos,
                    'total': pendentes + pagos,
                    'regime': regime
                })
                
            except Exception as e:
                logger.warning(f"Error parsing card: {e}")
                continue
        
        browser.close()
    
    # Sort by pendentes descending (largest first for optimal worker usage)
    entities.sort(key=lambda x: x['precatorios_pendentes'], reverse=True)
    
    logger.info(f"✅ Loaded {len(entities)} entities")
    return entities


def clean_ordem_column(df: pd.DataFrame) -> pd.DataFrame:
    """Convert ordem column to numeric (remove ordinal suffixes like º, °, ª)"""
    if 'ordem' in df.columns:
        try:
            df['ordem'] = df['ordem'].astype(str).str.replace(r'[º°ª]', '', regex=True)
            df['ordem'] = pd.to_numeric(df['ordem'], errors='coerce').fillna(0).astype(int)
            logger.info("✅ Converted 'ordem' to numeric")
        except Exception as e:
            logger.warning(f"⚠️ Could not convert ordem: {e}")
    return df


def format_monetary_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Convert monetary columns to numeric format"""
    monetary_cols = [
        'valor_requisitado', 'valor_pago', 'valor_prioridade', 
        'valor_rpv', 'valor_total', 'valor'
    ]
    
    for col in monetary_cols:
        if col in df.columns:
            try:
                df[col] = df[col].astype(str).str.replace('.', '', regex=False)
                df[col] = df[col].str.replace(',', '.', regex=False)
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
            except Exception as e:
                logger.warning(f"⚠️ Could not convert {col}: {e}")
    
    return df


def save_dataframe(df: pd.DataFrame, output_path: str, sheet_name: str = "Precatórios"):
    """Save DataFrame to CSV and Excel with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output_path = Path(output_path)
    output_path.parent.mkdir(parents=True, exist_ok=True)
    
    # Save CSV
    df.to_csv(output_path, index=False, encoding='utf-8-sig', sep=';', decimal=',')
    logger.info(f"💾 Saved CSV: {output_path}")
    
    # Save Excel
    excel_path = output_path.with_suffix('.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(50, max(12, max_length + 2))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    logger.info(f"💾 Saved Excel: {excel_path}")
    
    return excel_path


def main():
    parser = argparse.ArgumentParser(description='TJRJ Precatórios Scraper V5 - All Entities')
    
    parser.add_argument('--regime', choices=['geral', 'especial'], default='especial',
                       help='Regime to extract')
    parser.add_argument('--num-processes', type=int, default=12,
                       help='Number of parallel processes per entity')
    parser.add_argument('--timeout', type=int, default=60,
                       help='Timeout per entity (minutes)')
    parser.add_argument('--output', type=str,
                       help='Output CSV path (auto-generated if not specified)')
    parser.add_argument('--no-headless', action='store_true',
                       help='Show browser windows')
    parser.add_argument('--log-level', default='INFO',
                       help='Log level')
    parser.add_argument('--log-file', type=str,
                       help='Additional log file path')
    parser.add_argument('--entity-ids', type=str,
                       help='Comma-separated list of entity IDs to process (optional)')
    parser.add_argument('--skip-entity-ids', type=str,
                       help='Comma-separated list of entity IDs to skip (optional)')
    
    args = parser.parse_args()
    
    # Setup signal handlers
    signal.signal(signal.SIGINT, signal_handler)
    signal.signal(signal.SIGTERM, signal_handler)
    
    # Setup logging
    log_path = setup_logging(args.log_level, args.log_file)
    print(f"LOG_FILE:{log_path}")
    
    logger.info(f"\n{'='*80}")
    logger.info(f"🚀 TJRJ Precatórios Scraper V5 - All Entities")
    logger.info(f"{'='*80}")
    logger.info(f"Regime: {args.regime}")
    logger.info(f"Workers: {args.num_processes}")
    logger.info(f"Timeout: {args.timeout} min/entity")
    
    headless = not args.no_headless
    
    # Load entities from website
    entities = load_entities_from_website(args.regime, headless=headless)
    
    if not entities:
        logger.error("No entities found!")
        return 1
    
    # Filter entities if specified
    if args.entity_ids:
        target_ids = set(int(x.strip()) for x in args.entity_ids.split(','))
        entities = [e for e in entities if e['id'] in target_ids]
        logger.info(f"Filtered to {len(entities)} entities by ID")
    
    if args.skip_entity_ids:
        skip_ids = set(int(x.strip()) for x in args.skip_entity_ids.split(','))
        entities = [e for e in entities if e['id'] not in skip_ids]
        logger.info(f"Skipping {len(skip_ids)} entities")
    
    # Calculate totals
    total_pendentes = sum(e['precatorios_pendentes'] for e in entities)
    total_pages = sum((e['precatorios_pendentes'] + 9) // 10 for e in entities)
    
    logger.info(f"\n📊 Summary:")
    logger.info(f"  Entities: {len(entities)}")
    logger.info(f"  Total pendentes: {total_pendentes:,}")
    logger.info(f"  Total pages: {total_pages:,}")
    
    # Show top 5 entities
    logger.info(f"\n🏛️ Top 5 entities by size:")
    for i, e in enumerate(entities[:5], 1):
        pages = (e['precatorios_pendentes'] + 9) // 10
        logger.info(f"  {i}. {e['nome']}: {e['precatorios_pendentes']:,} pendentes ({pages:,} pages)")
    
    # === MAIN EXTRACTION LOOP ===
    all_records = []
    entity_stats = []
    start_time = time.time()
    
    for idx, entity in enumerate(entities, 1):
        if SHUTDOWN_REQUESTED:
            logger.warning("⚠️ Shutdown requested - stopping before next entity")
            break
        
        entity_pages = (entity['precatorios_pendentes'] + 9) // 10
        
        logger.info(f"\n{'='*80}")
        logger.info(f"📍 ENTITY {idx}/{len(entities)}: {entity['nome']}")
        logger.info(f"{'='*80}")
        
        # Calculate dynamic timeout based on pages
        # ~3 seconds per page + 10 min margin
        dynamic_timeout = max(args.timeout, (entity_pages * 3) // 60 + 10)
        
        records, stats = extract_single_entity(
            entity_id=entity['id'],
            entity_name=entity['nome'],
            regime=args.regime,
            total_pages=entity_pages,
            num_processes=args.num_processes,
            headless=headless,
            timeout_minutes=dynamic_timeout
        )
        
        all_records.extend(records)
        entity_stats.append(stats)
        
        # Progress update
        elapsed = time.time() - start_time
        logger.info(f"\n📈 Progress: {idx}/{len(entities)} entities | {len(all_records):,} total records | {elapsed/60:.1f}min elapsed")
    
    # === FINAL DATA PROCESSING ===
    logger.info(f"\n{'='*80}")
    logger.info(f"📦 FINAL PROCESSING")
    logger.info(f"{'='*80}")
    logger.info(f"Total records collected: {len(all_records):,}")
    
    if all_records:
        df = pd.DataFrame(all_records)
        
        # Clean and format
        df = clean_ordem_column(df)
        df = format_monetary_columns(df)
        
        # Sort by entidade_devedora and ordem
        sort_cols = []
        if 'entidade_devedora' in df.columns:
            sort_cols.append('entidade_devedora')
        if 'ordem' in df.columns:
            sort_cols.append('ordem')
        
        if sort_cols:
            df = df.sort_values(sort_cols)
            logger.info(f"📊 Sorted by: {', '.join(sort_cols)}")
        
        # Generate output path
        timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
        output_path = args.output or f"output/precatorios_{args.regime}_ALL_{timestamp}.csv"
        
        # Save ONCE at the end
        save_dataframe(df, output_path)
        
        logger.info(f"\n✅ EXTRACTION COMPLETE!")
        logger.info(f"  Total records: {len(df):,}")
        logger.info(f"  Entities processed: {len(entity_stats)}")
        logger.info(f"  Total time: {(time.time() - start_time)/60:.1f} min")
        logger.info(f"  Output: {output_path}")
    else:
        logger.warning("⚠️ No records extracted!")
        return 1
    
    # Summary stats
    successful = sum(1 for s in entity_stats if s.get('success', False))
    failed = len(entity_stats) - successful
    
    if failed > 0:
        logger.warning(f"⚠️ {failed} entities had issues")
        return 1
    
    return 0


if __name__ == "__main__":
    sys.exit(main())
