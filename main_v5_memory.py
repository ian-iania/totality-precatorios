"""
TJRJ Precatórios Scraper V5 - Full Memory Mode

Key improvements:
- ALL data accumulated in memory - NO intermediate I/O
- Single worker per entity (simpler, more reliable)
- Validation: expected records vs actual records
- Sorted by (entidade_devedora, ordem) at the end
- Excel output with auto-filter, styled headers, freeze panes

Usage:
    python main_v5_memory.py --entities-json '[{"id":1,"name":"...","pages":10},...]' --regime especial_rj
"""

import sys
import time
import json
import argparse
import multiprocessing as mp
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Tuple
from loguru import logger
import pandas as pd

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from playwright.sync_api import sync_playwright
from src.scraper_v3 import TJRJPrecatoriosScraperV3
from src.models import ScraperConfig


class SimpleEntity:
    """Simple entity object compatible with scraper expectations"""
    def __init__(self, id: int, name: str, regime: str):
        self.id_entidade = id
        self.nome_entidade = name
        self.regime = regime


def setup_logging(level: str = "INFO"):
    """Configure logging"""
    logger.remove()
    
    # Console output
    logger.add(
        sys.stderr,
        format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <7}</level> | <level>{message}</level>",
        level=level,
        colorize=True
    )
    
    # File output - single file for all
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    logger.add(
        log_dir / "scraper_v3.log",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <7} | {message}",
        level=level,
        rotation="10 MB"
    )


def extract_page_range_worker(args: Dict) -> Dict:
    """
    Worker function to extract a RANGE of pages from an entity.
    Used for both full entity extraction and partial (hybrid mode).
    Returns all records in memory.
    """
    entity_id = args['entity_id']
    entity_name = args['entity_name']
    regime = args['regime']
    start_page = args.get('start_page', 1)
    end_page = args.get('end_page', args.get('total_pages', 1))
    worker_id = args.get('worker_id', 0)
    headless = args.get('headless', True)
    timeout_minutes = args.get('timeout_minutes', 30)
    total_pages = end_page - start_page + 1
    
    start_time = time.time()
    timeout_seconds = timeout_minutes * 60
    
    all_records = []
    
    try:
        # Initialize scraper
        config = ScraperConfig(
            headless=headless,
            max_retries=3,
            use_cache=True,
            skip_expanded_details=True
        )
        scraper = TJRJPrecatoriosScraperV3(config=config, skip_expanded=True)
        
        # Create simple entity object compatible with scraper
        entidade = SimpleEntity(id=entity_id, name=entity_name, regime=regime)
        
        logger.info(f"[E{entity_id}:W{worker_id}] 🌐 Starting {entity_name} pages {start_page}-{end_page} ({total_pages} pages)")
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            context = browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/120.0.0.0 Safari/537.36'
            )
            page = context.new_page()
            
            # Navigate to entity page
            url = f"https://www3.tjrj.jus.br/PortalConhecimento/precatorio#!/ordem-cronologica?idEntidadeDevedora={entity_id}"
            page.goto(url, wait_until='networkidle', timeout=60000)
            
            # Wait for table
            try:
                page.wait_for_selector("text=/Número.*Precatório/i", timeout=15000)
            except:
                pass
            page.wait_for_timeout(2000)
            
            logger.info(f"[E{entity_id}:W{worker_id}] ✅ Page loaded")
            
            # Navigate to start page if not page 1
            if start_page > 1:
                logger.info(f"[E{entity_id}:W{worker_id}] ⏩ Navigating to page {start_page}...")
                for nav_page in range(1, start_page):
                    next_selectors = [
                        'a[ng-click="vm.ProximaPagina()"]',
                        'text=Próxima',
                        'a:has-text("Próxima")',
                    ]
                    try:
                        page.wait_for_selector('.block-ui-overlay', state='hidden', timeout=3000)
                    except:
                        pass
                    for selector in next_selectors:
                        try:
                            btn = page.query_selector(selector)
                            if btn and btn.is_visible():
                                btn.click()
                                page.wait_for_timeout(800)
                                break
                        except:
                            continue
                    if nav_page % 50 == 0:
                        logger.info(f"[E{entity_id}:W{worker_id}] ⏩ Navigation progress: {nav_page}/{start_page-1}")
                logger.info(f"[E{entity_id}:W{worker_id}] ✅ Reached page {start_page}")
            
            # Extract page range
            current_page = start_page
            while current_page <= end_page:
                # Check timeout
                if time.time() - start_time > timeout_seconds:
                    logger.warning(f"[E{entity_id}] ⏰ Timeout after {timeout_minutes}min")
                    break
                
                # Extract current page
                precatorios = scraper._extract_precatorios_from_page(page, entidade)
                
                for prec in precatorios:
                    all_records.append(prec.model_dump())
                
                if current_page % 10 == 0 or current_page == end_page:
                    logger.info(f"[E{entity_id}:W{worker_id}] Page {current_page}/{end_page} - {len(all_records)} records")
                
                # Navigate to next page if not last
                if current_page < end_page:
                    next_selectors = [
                        'a[ng-click="vm.ProximaPagina()"]',
                        'text=Próxima',
                        'a:has-text("Próxima")',
                    ]
                    
                    # Wait for overlay
                    try:
                        page.wait_for_selector('.block-ui-overlay', state='hidden', timeout=3000)
                    except:
                        pass
                    
                    clicked = False
                    for selector in next_selectors:
                        try:
                            btn = page.query_selector(selector)
                            if btn and btn.is_visible():
                                btn.click()
                                page.wait_for_timeout(1500)
                                clicked = True
                                break
                        except:
                            continue
                    
                    if not clicked:
                        logger.warning(f"[E{entity_id}:W{worker_id}] ⚠️ Next button not found on page {current_page}")
                        break
                
                current_page += 1
            
            # Close browser
            try:
                context.close()
                browser.close()
            except:
                pass
        
        elapsed = time.time() - start_time
        logger.info(f"[E{entity_id}:W{worker_id}] ✅ Complete: {len(all_records)} records in {elapsed/60:.1f}min")
        
        return {
            'entity_id': entity_id,
            'entity_name': entity_name,
            'worker_id': worker_id,
            'start_page': start_page,
            'end_page': end_page,
            'records': all_records,
            'records_count': len(all_records),
            'expected_records': total_pages * 10,
            'elapsed_seconds': elapsed,
            'success': True,
            'error': None
        }
        
    except Exception as e:
        elapsed = time.time() - start_time
        logger.error(f"[E{entity_id}:W{worker_id}] ❌ Failed: {e}")
        return {
            'entity_id': entity_id,
            'entity_name': entity_name,
            'worker_id': worker_id,
            'start_page': start_page,
            'end_page': end_page,
            'records': all_records,
            'records_count': len(all_records),
            'expected_records': total_pages * 10,
            'elapsed_seconds': elapsed,
            'success': False,
            'error': str(e)
        }


def run_full_extraction(
    entities: List[Dict],
    regime: str,
    max_concurrent: int = 10,
    headless: bool = True,
    timeout_minutes: int = 30,
    output_path: str = None
) -> Tuple[pd.DataFrame, Dict]:
    """
    Extract ALL entities, accumulate in memory, write CSV once at the end.
    Simple mode: 1 worker per entity.
    
    Args:
        entities: List of {"id": int, "name": str, "pages": int}
        regime: 'geral' or 'especial'
        max_concurrent: Max concurrent workers (default 10)
        headless: Run browsers headless
        timeout_minutes: Timeout per worker
        output_path: Final CSV path
    
    Returns:
        (DataFrame, stats)
    """
    start_time = time.time()
    
    # Calculate totals for validation
    total_expected_records = sum(e['pages'] * 10 for e in entities)
    total_pages = sum(e['pages'] for e in entities)
    
    # Prepare worker args - 1 worker per entity
    worker_args = []
    for entity in entities:
        worker_args.append({
            'entity_id': entity['id'],
            'entity_name': entity['name'],
            'regime': regime,
            'start_page': 1,
            'end_page': entity['pages'],
            'worker_id': 0,
            'headless': headless,
            'timeout_minutes': timeout_minutes
        })
    
    logger.info(f"\n{'='*80}")
    logger.info(f"🚀 V5 FULL MEMORY EXTRACTION - {regime.upper()}")
    logger.info(f"{'='*80}")
    logger.info(f"Entities: {len(entities)}")
    logger.info(f"Total pages: {total_pages:,}")
    logger.info(f"Expected records: {total_expected_records:,}")
    logger.info(f"Workers: {len(worker_args)} (1 per entity), {max_concurrent} concurrent")
    logger.info(f"{'='*80}\n")
    
    # Accumulator for all records
    all_records = []
    results = []
    
    # Process with limited concurrency
    ctx = mp.get_context('spawn')
    
    with ctx.Pool(processes=max_concurrent) as pool:
        # Submit all workers
        async_results = []
        for args in worker_args:
            task_id = f"E{args['entity_id']}"
            async_results.append((task_id, args, pool.apply_async(extract_page_range_worker, (args,))))
        
        # Collect results with polling (avoid deadlocks)
        pending = list(async_results)
        
        while pending:
            still_pending = []
            for task_id, args, async_result in pending:
                try:
                    result = async_result.get(timeout=1.0)
                    results.append(result)
                    
                    if result['success']:
                        all_records.extend(result['records'])
                        logger.info(f"✅ {task_id} {result['entity_name']}: {result['records_count']} records")
                    else:
                        logger.error(f"❌ {task_id} {result['entity_name']}: {result['error']}")
                        # Still add partial records
                        all_records.extend(result['records'])
                        
                except mp.TimeoutError:
                    still_pending.append((task_id, args, async_result))
                except Exception as e:
                    logger.error(f"❌ {task_id} ERROR: {e}")
                    results.append({
                        'entity_id': args['entity_id'],
                        'entity_name': args['entity_name'],
                        'records': [],
                        'records_count': 0,
                        'success': False,
                        'error': str(e)
                    })
            
            pending = still_pending
            if pending:
                # Log progress
                done = len(worker_args) - len(pending)
                logger.info(f"⏳ Progress: {done}/{len(worker_args)} entities complete, {len(all_records):,} records in memory")
                time.sleep(5)  # Check every 5 seconds
    
    # === VALIDATION ===
    actual_records = len(all_records)
    logger.info(f"\n{'='*80}")
    logger.info(f"📊 VALIDATION")
    logger.info(f"{'='*80}")
    logger.info(f"Expected records: {total_expected_records:,}")
    logger.info(f"Actual records:   {actual_records:,}")
    
    if actual_records == total_expected_records:
        logger.info(f"✅ PERFECT MATCH!")
    elif actual_records > total_expected_records * 0.95:
        logger.warning(f"⚠️ Minor difference: {total_expected_records - actual_records} records missing ({100*actual_records/total_expected_records:.1f}%)")
    else:
        logger.error(f"❌ SIGNIFICANT DIFFERENCE: {total_expected_records - actual_records} records missing ({100*actual_records/total_expected_records:.1f}%)")
    
    # === HELPER FUNCTION FOR SAVING ===
    def save_dataframe(df: pd.DataFrame, csv_path: str, sheet_name: str = "Precatórios"):
        """Save DataFrame to CSV and Excel with formatting"""
        if df.empty:
            return
        
        # Clean 'ordem' column
        if 'ordem' in df.columns:
            try:
                df['ordem'] = df['ordem'].astype(str).str.replace(r'[º°ª]', '', regex=True).str.strip()
                df['ordem'] = pd.to_numeric(df['ordem'], errors='coerce').fillna(0).astype(int)
            except:
                pass
        
        # Format monetary columns
        for col in ['valor_historico', 'saldo_atualizado']:
            if col in df.columns:
                try:
                    df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0).round(2)
                except:
                    pass
        
        # Sort
        if 'entidade_devedora' in df.columns and 'ordem' in df.columns:
            try:
                df = df.sort_values(['entidade_devedora', 'ordem'])
            except:
                pass
        
        # Save CSV
        Path(csv_path).parent.mkdir(parents=True, exist_ok=True)
        df.to_csv(csv_path, index=False, encoding='utf-8-sig', sep=';', decimal=',')
        logger.info(f"💾 Saved CSV: {csv_path} ({len(df):,} records)")
        
        # Save Excel
        try:
            from openpyxl import Workbook
            from openpyxl.utils.dataframe import dataframe_to_rows
            from openpyxl.styles import Font, Alignment, PatternFill
            from openpyxl.utils import get_column_letter
            
            excel_path = csv_path.replace('.csv', '.xlsx')
            wb = Workbook()
            ws = wb.active
            ws.title = sheet_name
            
            for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
                for c_idx, value in enumerate(row, 1):
                    cell = ws.cell(row=r_idx, column=c_idx, value=value)
                    if r_idx == 1:
                        cell.font = Font(bold=True, color="FFFFFF")
                        cell.fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
                        cell.alignment = Alignment(horizontal="center", vertical="center")
            
            ws.auto_filter.ref = ws.dimensions
            ws.freeze_panes = "A2"
            
            for col_idx, column in enumerate(df.columns, 1):
                max_length = len(str(column))
                for row in ws.iter_rows(min_row=2, min_col=col_idx, max_col=col_idx):
                    for cell in row:
                        if cell.value:
                            max_length = max(max_length, len(str(cell.value)))
                ws.column_dimensions[get_column_letter(col_idx)].width = min(max_length + 2, 50)
            
            wb.save(excel_path)
            logger.info(f"📊 Saved Excel: {excel_path}")
        except Exception as e:
            logger.warning(f"⚠️ Failed to generate Excel: {e}")
    
    # === SAVE FILE ===
    logger.info(f"\n📦 Saving file...")
    
    # Create DataFrame
    df = pd.DataFrame(all_records) if all_records else pd.DataFrame()
    
    # Save to output path
    if output_path and not df.empty:
        save_dataframe(df, output_path, "Precatórios")
    
    elapsed = time.time() - start_time
    
    stats = {
        'total_entities': len(entities),
        'successful_entities': sum(1 for r in results if r['success']),
        'failed_entities': sum(1 for r in results if not r['success']),
        'expected_records': total_expected_records,
        'actual_records': actual_records,
        'match_percent': 100 * actual_records / total_expected_records if total_expected_records > 0 else 0,
        'elapsed_seconds': elapsed,
        'records_per_second': actual_records / elapsed if elapsed > 0 else 0
    }
    
    logger.info(f"\n{'='*80}")
    logger.info(f"✅ EXTRACTION COMPLETE!")
    logger.info(f"{'='*80}")
    logger.info(f"Entities: {stats['successful_entities']}/{stats['total_entities']} successful")
    logger.info(f"Records: {stats['actual_records']:,} ({stats['match_percent']:.1f}% of expected)")
    logger.info(f"Time: {elapsed/60:.1f} min ({stats['records_per_second']:.1f} rec/s)")
    logger.info(f"{'='*80}\n")
    
    return df, stats


def main():
    parser = argparse.ArgumentParser(description='TJRJ Precatórios Scraper V5 - Full Memory Mode')
    
    parser.add_argument('--entities-json', type=str, required=True, 
                       help='JSON array of entities: [{"id":1,"name":"...","pages":10},...]')
    parser.add_argument('--regime', choices=['geral', 'especial'], required=True)
    parser.add_argument('--max-concurrent', type=int, default=4, help='Max concurrent entity extractions')
    parser.add_argument('--timeout', type=int, default=30, help='Timeout per entity (minutes)')
    parser.add_argument('--output', type=str, help='Output CSV path')
    parser.add_argument('--no-headless', action='store_true', help='Show browser')
    parser.add_argument('--log-level', default='INFO', help='Log level')
    
    args = parser.parse_args()
    
    setup_logging(args.log_level)
    
    # Parse entities
    try:
        entities = json.loads(args.entities_json)
    except json.JSONDecodeError as e:
        logger.error(f"Invalid JSON: {e}")
        return 1
    
    # Filter entities with pages > 0
    entities = [e for e in entities if e.get('pages', 0) > 0]
    
    if not entities:
        logger.error("No entities to process")
        return 1
    
    # Generate output path
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = args.output or f"output/precatorios_{args.regime}_{timestamp}.csv"
    
    # Run extraction
    df, stats = run_full_extraction(
        entities=entities,
        regime=args.regime,
        max_concurrent=args.max_concurrent,
        headless=not args.no_headless,
        timeout_minutes=args.timeout,
        output_path=output_path
    )
    
    return 0 if stats['failed_entities'] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
