"""
TJRJ Precatórios Scraper V4 Memory - Full In-Memory Mode

Based on V4 Fast but with:
- NO intermediate I/O (no partial files)
- All data accumulated in memory until final save
- Data formatting: ordem as numeric, monetary columns as numbers
- Sorted by entidade_devedora and ordem
- Excel output with auto-filter, styled headers, freeze panes

Usage:
    python main_v4_memory.py --entity-id 1 --total-pages 2984 --num-processes 4
"""

import sys
import time
import argparse
import multiprocessing as mp
import re
from pathlib import Path
from datetime import datetime
from typing import List, Dict, Optional, Tuple
from loguru import logger
import pandas as pd

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent))

from src.scraper_v3 import TJRJPrecatoriosScraperV3
from src.models import ScraperConfig, EntidadeDevedora


def setup_logging(level: str = "INFO"):
    """Configure logging"""
    logger.remove()
    
    # Console output
    logger.add(
        sys.stderr,
        format="<green>{time:YYYY-MM-DD HH:mm:ss}</green> | <level>{level: <7}</level> | <level>{message}</level>",
        level=level,
        colorize=True
    )
    
    # File output
    log_dir = Path("logs")
    log_dir.mkdir(exist_ok=True)
    
    logger.add(
        log_dir / "scraper_v3.log",
        format="{time:YYYY-MM-DD HH:mm:ss} | {level: <7} | {message}",
        level="DEBUG",
        rotation="10 MB"
    )


def divide_pages_into_ranges(total_pages: int, num_processes: int) -> List[Tuple[int, int]]:
    """Divide total pages into ranges for parallel processing"""
    pages_per_process = total_pages // num_processes
    remainder = total_pages % num_processes
    
    ranges = []
    start = 1
    
    for i in range(num_processes):
        extra = 1 if i < remainder else 0
        end = start + pages_per_process - 1 + extra
        ranges.append((start, end))
        start = end + 1
    
    return ranges


def extract_worker(args: Dict) -> Dict:
    """
    Worker function for parallel extraction
    
    Returns dict with results directly in memory (NO disk I/O)
    Maintains same log format for UI compatibility
    """
    from playwright.sync_api import sync_playwright
    
    entity_id = args['entity_id']
    entity_name = args['entity_name']
    regime = args['regime']
    start_page = args['start_page']
    end_page = args['end_page']
    process_id = args['process_id']
    skip_expanded = args.get('skip_expanded', True)
    headless = args.get('headless', True)
    timeout_minutes = args.get('timeout_minutes', 30)
    
    start_time = time.time()
    timeout_seconds = timeout_minutes * 60
    
    # Accumulate records in memory only
    precatorios_data = []
    
    try:
        logger.info(f"[P{process_id}] 🚀 Starting: pages {start_page}-{end_page} (timeout: {timeout_minutes}min)")
        
        # Create scraper
        config = ScraperConfig(headless=headless)
        scraper = TJRJPrecatoriosScraperV3(config=config, skip_expanded=skip_expanded)
        
        # Create entity
        entidade = EntidadeDevedora(
            id_entidade=entity_id,
            nome_entidade=entity_name,
            regime=regime,
            precatorios_pagos=0,
            precatorios_pendentes=0,
            valor_prioridade=0,
            valor_rpv=0
        )
        
        with sync_playwright() as p:
            browser = p.chromium.launch(headless=headless)
            context = browser.new_context(
                viewport={'width': 1920, 'height': 1080},
                user_agent='Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) '
                          'AppleWebKit/537.36 (KHTML, like Gecko) '
                          'Chrome/120.0.0.0 Safari/537.36'
            )
            page = context.new_page()
            
            # Navigate to entity
            logger.info(f"[P{process_id}] 🌐 Navigating to entity page...")
            entity_url = f"https://www3.tjrj.jus.br/PortalConhecimento/precatorio#!/ordem-cronologica?idEntidadeDevedora={entity_id}"
            page.goto(entity_url, wait_until='networkidle', timeout=60000)
            page.wait_for_timeout(3000)
            logger.info(f"[P{process_id}] ✅ Entity page loaded")
            
            # Go to start page
            if start_page > 1:
                logger.info(f"[P{process_id}] 🔄 Jumping to start page {start_page}...")
                if not scraper.goto_page_direct(page, start_page):
                    raise Exception(f"Failed to navigate to page {start_page}")
                logger.info(f"[P{process_id}] ✅ Arrived at page {start_page}")
            
            # Extract pages
            current_page = start_page
            total_pages_in_range = end_page - start_page + 1
            
            while current_page <= end_page:
                # Check timeout
                elapsed = time.time() - start_time
                if elapsed > timeout_seconds:
                    logger.warning(f"[P{process_id}] Timeout after {elapsed/60:.1f}min")
                    break
                
                page_in_range = current_page - start_page + 1
                
                # Heartbeat logging every 50 pages
                if page_in_range % 50 == 0:
                    speed = len(precatorios_data) / elapsed if elapsed > 0 else 0
                    logger.info(f"[P{process_id}] 💓 Heartbeat: {page_in_range}/{total_pages_in_range} pages, {len(precatorios_data)} records, {speed:.1f} rec/s")
                
                # Log format compatible with UI: [P1] Page X/Y (Z/W)
                logger.info(f"[P{process_id}] Page {current_page}/{end_page} ({page_in_range}/{total_pages_in_range})")
                
                # Extract precatórios from current page
                precatorios = scraper._extract_precatorios_from_page(page, entidade)
                
                # Convert to dicts and accumulate in memory
                for prec in precatorios:
                    precatorios_data.append(prec.model_dump())
                
                # Log format compatible with UI: [P1] ✅ ... (total: N)
                logger.info(f"[P{process_id}]   ✅ {len(precatorios)} records (total: {len(precatorios_data)})")
                
                # Next page navigation
                if current_page < end_page:
                    pages_remaining = end_page - current_page
                    
                    if pages_remaining <= 5:
                        # Robust navigation for final pages
                        next_clicked = False
                        for retry in range(3):
                            try:
                                page.wait_for_selector('.block-ui-overlay', state='hidden', timeout=3000)
                            except:
                                pass
                            
                            next_btn = page.query_selector('a[ng-click="vm.ProximaPagina()"]')
                            if next_btn and next_btn.is_visible():
                                try:
                                    next_btn.click()
                                    page.wait_for_timeout(2000)
                                    next_clicked = True
                                    break
                                except Exception as click_err:
                                    logger.warning(f"[P{process_id}] ⚠️ Click failed (retry {retry+1}): {click_err}")
                                    page.wait_for_timeout(1000)
                            else:
                                page.wait_for_timeout(2000)
                        
                        if not next_clicked:
                            logger.warning(f"[P{process_id}] ⚠️ Failed to navigate after 3 retries on page {current_page}")
                    else:
                        # Fast navigation for most pages
                        next_btn = page.query_selector('a[ng-click="vm.ProximaPagina()"]')
                        if next_btn:
                            next_btn.click()
                            page.wait_for_timeout(2000)
                
                current_page += 1
            
            # Close browser
            logger.info(f"[P{process_id}] 🔄 Extraction complete, closing browser...")
            try:
                context.close()
                browser.close()
                logger.info(f"[P{process_id}] ✅ Browser closed")
            except Exception as close_error:
                logger.warning(f"[P{process_id}] ⚠️ Browser close issue: {close_error}")
        
        elapsed = time.time() - start_time
        logger.info(f"[P{process_id}] ✅ Complete: {len(precatorios_data)} records in {elapsed/60:.1f}min")
        
        # Return data in memory only - NO disk I/O
        return {
            'process_id': process_id,
            'entity_id': entity_id,
            'entity_name': entity_name,
            'start_page': start_page,
            'end_page': end_page,
            'records': precatorios_data,
            'records_count': len(precatorios_data),
            'elapsed_seconds': elapsed,
            'success': True,
            'error': None
        }
    
    except Exception as e:
        elapsed = time.time() - start_time
        logger.error(f"[P{process_id}] ❌ Failed: {e}")
        
        # Return whatever we have in memory
        return {
            'process_id': process_id,
            'entity_id': entity_id,
            'entity_name': entity_name,
            'start_page': start_page,
            'end_page': end_page,
            'records': precatorios_data,
            'records_count': len(precatorios_data),
            'elapsed_seconds': elapsed,
            'success': False,
            'error': str(e)
        }


def clean_ordem_column(df: pd.DataFrame) -> pd.DataFrame:
    """Convert ordem column to numeric (remove ordinal suffixes like º, °, ª)"""
    if 'ordem' in df.columns:
        try:
            # Remove ordinal suffixes and convert to int
            df['ordem'] = df['ordem'].astype(str).str.replace(r'[º°ª]', '', regex=True)
            df['ordem'] = pd.to_numeric(df['ordem'], errors='coerce').fillna(0).astype(int)
            logger.info("✅ Converted 'ordem' to numeric")
        except Exception as e:
            logger.warning(f"⚠️ Could not convert ordem: {e}")
    return df


def format_monetary_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Convert monetary columns to numeric format"""
    monetary_cols = [
        'valor_requisitado', 'valor_pago', 'valor_prioridade', 
        'valor_rpv', 'valor_total', 'valor'
    ]
    
    for col in monetary_cols:
        if col in df.columns:
            try:
                # Handle Brazilian format: 1.234,56 -> 1234.56
                df[col] = df[col].astype(str).str.replace('.', '', regex=False)
                df[col] = df[col].str.replace(',', '.', regex=False)
                df[col] = pd.to_numeric(df[col], errors='coerce').fillna(0)
                logger.info(f"✅ Converted '{col}' to numeric")
            except Exception as e:
                logger.warning(f"⚠️ Could not convert {col}: {e}")
    
    return df


def save_dataframe(df: pd.DataFrame, output_path: str, sheet_name: str = "Precatórios"):
    """Save DataFrame to CSV and Excel with formatting"""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    output_path = Path(output_path)
    
    # Save CSV
    df.to_csv(output_path, index=False, encoding='utf-8-sig', sep=';', decimal=',')
    logger.info(f"💾 Saved CSV: {output_path}")
    
    # Save Excel
    excel_path = output_path.with_suffix('.xlsx')
    
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name
    
    # Write data
    for r_idx, row in enumerate(dataframe_to_rows(df, index=False, header=True), 1):
        for c_idx, value in enumerate(row, 1):
            ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Style header row
    header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
    header_font = Font(bold=True, color="FFFFFF")
    
    for cell in ws[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal='center')
    
    # Freeze header row
    ws.freeze_panes = 'A2'
    
    # Auto-filter
    ws.auto_filter.ref = ws.dimensions
    
    # Adjust column widths
    for column in ws.columns:
        max_length = 0
        column_letter = column[0].column_letter
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = min(50, max(12, max_length + 2))
        ws.column_dimensions[column_letter].width = adjusted_width
    
    wb.save(excel_path)
    logger.info(f"💾 Saved Excel: {excel_path}")
    
    return excel_path


def run_parallel_extraction(
    entity_id: int,
    entity_name: str,
    regime: str,
    total_pages: int,
    num_processes: int = 4,
    skip_expanded: bool = True,
    headless: bool = True,
    timeout_minutes: int = 30,
    output_path: Optional[str] = None,
    append_mode: bool = False
) -> Tuple[pd.DataFrame, Dict]:
    """
    Run parallel extraction with full in-memory handling
    
    NO intermediate I/O - all data stays in memory until final save
    """
    start_time = time.time()
    
    # Divide pages
    ranges = divide_pages_into_ranges(total_pages, num_processes)
    
    logger.info(f"\n{'='*80}")
    logger.info(f"🚀 V4 Memory Extraction - {entity_name}")
    logger.info(f"{'='*80}")
    logger.info(f"Pages: {total_pages} | Processes: {num_processes} | Skip expanded: {skip_expanded}")
    logger.info(f"Mode: FULL IN-MEMORY (no intermediate I/O)")
    
    for i, (start, end) in enumerate(ranges, 1):
        logger.info(f"  P{i}: pages {start:,}-{end:,} ({end-start+1:,} pages)")
    
    # Prepare worker args
    worker_args = []
    for i, (start, end) in enumerate(ranges, 1):
        worker_args.append({
            'entity_id': entity_id,
            'entity_name': entity_name,
            'regime': regime,
            'start_page': start,
            'end_page': end,
            'process_id': i,
            'skip_expanded': skip_expanded,
            'headless': headless,
            'timeout_minutes': timeout_minutes
        })
    
    # Accumulate all records in memory
    all_records = []
    results = []
    
    worker_timeout = timeout_minutes * 60
    
    logger.info(f"\n🔄 Starting extraction (worker timeout: {worker_timeout}s = {timeout_minutes}min)...")
    
    with mp.Pool(processes=num_processes) as pool:
        # Submit all workers
        async_results = []
        for args in worker_args:
            async_results.append((args['process_id'], pool.apply_async(extract_worker, (args,))))
        
        # Collect results with timeout
        for process_id, async_result in async_results:
            try:
                result = async_result.get(timeout=worker_timeout)
                results.append(result)
                
                if result['success']:
                    all_records.extend(result['records'])
                    logger.info(f"✅ P{result['process_id']} done: {result['records_count']} records")
                else:
                    logger.error(f"❌ P{result['process_id']} failed: {result['error']}")
                    # Still add partial records
                    all_records.extend(result['records'])
                    
            except mp.TimeoutError:
                logger.error(f"❌ P{process_id} TIMEOUT after {worker_timeout}s")
                results.append({
                    'process_id': process_id,
                    'records': [],
                    'records_count': 0,
                    'success': False,
                    'error': f'Timeout after {worker_timeout}s'
                })
            except Exception as e:
                logger.error(f"❌ P{process_id} ERROR: {e}")
                results.append({
                    'process_id': process_id,
                    'records': [],
                    'records_count': 0,
                    'success': False,
                    'error': str(e)
                })
    
    # === DATA PROCESSING ===
    logger.info(f"\n📦 Processing {len(all_records)} records...")
    
    if all_records:
        df = pd.DataFrame(all_records)
        
        # Clean ordem column (convert to numeric)
        df = clean_ordem_column(df)
        
        # Format monetary columns
        df = format_monetary_columns(df)
        
        # Sort by entidade_devedora and ordem
        sort_cols = []
        if 'entidade_devedora' in df.columns:
            sort_cols.append('entidade_devedora')
        if 'ordem' in df.columns:
            sort_cols.append('ordem')
        
        if sort_cols:
            df = df.sort_values(sort_cols)
            logger.info(f"📊 Sorted by: {', '.join(sort_cols)}")
    else:
        df = pd.DataFrame()
    
    # === SAVE FILES ===
    if output_path and not df.empty:
        if append_mode and Path(output_path).exists():
            # Read existing and append
            existing_df = pd.read_csv(output_path, sep=';', decimal=',', encoding='utf-8-sig')
            df = pd.concat([existing_df, df], ignore_index=True)
            logger.info(f"📎 Appended to existing file")
        
        # Save CSV and Excel
        save_dataframe(df, output_path)
    
    elapsed = time.time() - start_time
    
    stats = {
        'total_records': len(df),
        'elapsed_seconds': elapsed,
        'successful_processes': sum(1 for r in results if r['success']),
        'failed_processes': sum(1 for r in results if not r['success']),
        'records_per_second': len(df) / elapsed if elapsed > 0 else 0
    }
    
    logger.info(f"\n{'='*80}")
    logger.info(f"✅ COMPLETE!")
    logger.info(f"{'='*80}")
    logger.info(f"Records: {stats['total_records']:,}")
    logger.info(f"Time: {elapsed/60:.1f} min ({stats['records_per_second']:.1f} rec/s)")
    logger.info(f"Processes: {stats['successful_processes']}/{num_processes} successful")
    
    return df, stats


def main():
    parser = argparse.ArgumentParser(description='TJRJ Precatórios Scraper V4 Memory - Full In-Memory Mode')
    
    parser.add_argument('--entity-id', type=int, required=True, help='Entity ID')
    parser.add_argument('--entity-name', type=str, default='Unknown', help='Entity name')
    parser.add_argument('--regime', choices=['geral', 'especial'], default='especial')
    parser.add_argument('--total-pages', type=int, required=True, help='Total pages')
    parser.add_argument('--num-processes', type=int, default=4, help='Parallel processes')
    parser.add_argument('--timeout', type=int, default=30, help='Timeout per worker (minutes)')
    parser.add_argument('--output', type=str, help='Output CSV path')
    parser.add_argument('--append', action='store_true', help='Append to existing file')
    parser.add_argument('--no-headless', action='store_true', help='Show browser')
    parser.add_argument('--log-level', default='INFO', help='Log level')
    
    args = parser.parse_args()
    
    setup_logging(args.log_level)
    
    # Generate output path
    timestamp = datetime.now().strftime('%Y%m%d_%H%M%S')
    output_path = args.output or f"output/precatorios_regime_{args.regime}_{timestamp}.csv"
    
    # Run extraction
    df, stats = run_parallel_extraction(
        entity_id=args.entity_id,
        entity_name=args.entity_name,
        regime=args.regime,
        total_pages=args.total_pages,
        num_processes=args.num_processes,
        skip_expanded=True,
        headless=not args.no_headless,
        timeout_minutes=args.timeout,
        output_path=output_path,
        append_mode=args.append
    )
    
    return 0 if stats['failed_processes'] == 0 else 1


if __name__ == "__main__":
    sys.exit(main())
